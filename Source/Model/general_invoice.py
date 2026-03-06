"""!
********************************************************************************
@file   general_invoice.py
@brief  Generate invoice documents in XLSX and PDF with ZUGFeRD support.
********************************************************************************
"""

import os
import logging
from typing import Any
from io import BytesIO
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import qrcode

from PyQt6.QtWidgets import QMessageBox

from Source.version import __title__
from Source.Util.app_data import EXPORT_PATH, LOGO_ZUGFERD, EInvoiceOption
from Source.Util.openpyxl_util import XLSCreator
from Source.Model.data_handler import get_file_name, convert_xlsx_to_pdf, delete_file, \
    convert_to_de_date, convert_to_de_amount, convert_to_rate
from Source.Model.ZUGFeRD.drafthorse_data import INVOICE_TYPE, CURRENCY, \
    COUNTRY_CODE, PAYMENT_METHOD, VAT_CODE, UNIT
from Source.Model.ZUGFeRD.drafthorse_import import create_value_description
from Source.Model.ZUGFeRD.drafthorse_invoice import add_xml_to_pdf, create_factur_xml, convert_json_to_drafthorse_doc, fill_invoice_data

log = logging.getLogger(__title__)

FONT_NAME = "LiberationSans"
COLOR_LIGHT_GREY = "F2F2F2"
COLOR_LIGHT_BLUE = "CFF1FF"
COLOR_MIDDLE_GREY = "E0E0E0"
WHITE_BORDER = Border(right=Side(style="medium", color="FFFFFF"))
SECTION_KEY = "SECTION_KEY"

BOLD_SECTIONS = ["Fälliger Betrag"]
OBLIGATION_SECTIONS = ["Betrag (Netto)", "Steuerbetrag", "Steuersatz", "Summe Positionen (Netto)",
                       "Gesamt (Netto)", "Summe Umsatzsteuer", "Gesamt (Brutto)", "Fälliger Betrag"]


def write_data_to_excel(xls_creator: XLSCreator, ws: Worksheet, row: int, section_data: dict[str, str], title: str | None = None) -> int:
    """!
    @brief Write a labeled data section with key-value pairs to the Excel worksheet.
    @param xls_creator : XLS creator instance.
    @param ws : target worksheet.
    @param row : current row index.
    @param section_data : key-value pairs for the section.
    @param title : optional section title.
    @return next available row index.
    """
    data_present = False  # check to prevent that only title is written
    for description, content in section_data.items():
        if (content and (content != "0,00")) or (description in OBLIGATION_SECTIONS):
            data_present = True
            break
    if data_present:
        description_column = 1
        content_column = 4
        description_column_letter = get_column_letter(description_column)
        description_column_letter_end = get_column_letter(content_column - 1)
        content_column_letter = get_column_letter(content_column)
        content_column_letter_end = get_column_letter(9)
        if title is not None:
            xls_creator.set_cell(ws, row, description_column, title, fill_color=COLOR_LIGHT_BLUE, border=WHITE_BORDER, bold=True)
            xls_creator.set_cell(ws, row, content_column, fill_color=COLOR_LIGHT_BLUE, border=WHITE_BORDER, bold=True)
            ws.merge_cells(f"{description_column_letter}{row}:{description_column_letter_end}{row}")
            ws.merge_cells(f"{content_column_letter}{row}:{content_column_letter_end}{row}")
            row += 1
        count = 0
        last_section_name = ""
        for description, content in section_data.items():
            if (content and (content != "0,00")) or (description in OBLIGATION_SECTIONS):
                if content == SECTION_KEY:
                    last_section_name = description
                else:
                    if last_section_name:
                        fill_color: str | None = COLOR_MIDDLE_GREY
                        bold = True
                        xls_creator.set_cell(ws, row, description_column, last_section_name, fill_color=fill_color, border=WHITE_BORDER, bold=bold)
                        xls_creator.set_cell(ws, row, content_column, "", align="left", wrap_text=True, fill_color=fill_color, border=WHITE_BORDER)
                        ws.merge_cells(f"{description_column_letter}{row}:{description_column_letter_end}{row}")
                        ws.merge_cells(f"{content_column_letter}{row}:I{row}")
                        count = 0  # clear to start with white after section
                        row += 1
                        last_section_name = ""
                    fill_color = COLOR_LIGHT_GREY if (count % 2) else None
                    xls_creator.set_cell(ws, row, description_column, description, fill_color=fill_color, border=WHITE_BORDER)
                    bold = title == "Gesamtsummen" and description in BOLD_SECTIONS
                    xls_creator.set_cell(ws, row, content_column, content, align="left", wrap_text=True, fill_color=fill_color, border=WHITE_BORDER, bold=bold)
                    ws.merge_cells(f"{description_column_letter}{row}:{description_column_letter_end}{row}")
                    ws.merge_cells(f"{content_column_letter}{row}:I{row}")
                    count += 1
                    row += 1
        row += 1
    return row


def generate_epc_qr(invoice_data: dict[str, Any], box_size: int = 2, border: int = 0) -> BytesIO | None:
    """!
    @brief Generate an EPC QR code for SEPA credit transfer payment.
    @param invoice_data : invoice data as JSON dictionary.
    @param box_size : pixel size per QR module.
    @param border : border width in modules.
    @return PNG image buffer or None if no amount present.
    """
    data_payment = invoice_data["payment"]
    if data_payment["methods"]:
        payment_method = data_payment["methods"][0]
        name = payment_method["accountName"]  # Name des Zahlungskontos (BT-85)
        iban = payment_method["iban"]  # Kennung des Zahlungskontos (BT-84)
        bic = payment_method["bic"]  # Kennung des Zahlungsdienstleisters (BT-86)
    else:
        name = ""
        iban = ""
        bic = ""
    reference = data_payment["reference"]  # Verwendungszweck (BT-83)
    amount = invoice_data["totals"]["dueAmount"]  # Fälliger Betrag (BT-115)
    currency_code = invoice_data["currencyCode"]

    if amount:  # generate only for present amount
        # Build EPC string
        epc_data = [
            "BCD",                    # Service tag (fix)
            "002",                    # Version (001 (older) or 002)
            "1",                      # Character set (1 = UTF-8)
            "SCT",                    # Identification (SEPA Credit Transfer)
            bic,                      # BIC (optional, need in version 1)
            name,                     # Beneficiary name
            iban,                     # IBAN
            f"{currency_code}{amount:.2f}",  # Amount
            "",                       # Purpose (optional)
            reference,                # Structured reference
            ""                        # Unstructured remittance
        ]

        epc_string = "\n".join(epc_data)

        # Generate QR code
        qr = qrcode.QRCode(version=1, box_size=box_size, border=border)
        qr.add_data(epc_string)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        buffer = BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
    else:
        buffer = None
    return buffer


###########################
##     Create Invoice    ##
###########################

def convert_json_to_invoice(invoice_data: dict[str, Any], invoice_option: EInvoiceOption, create_qr_code: bool = False) -> None | str:
    """!
    @brief Convert JSON invoice data to XLSX/PDF/ZUGFeRD document.
    @param invoice_data : invoice data as JSON dictionary.
    @param invoice_option : output format option.
    @param create_qr_code : whether to include EPC QR code.
    @return file path to open or None on failure.
    """
    fill_invoice_data(invoice_data)

    needs_xml = invoice_option in [EInvoiceOption.XML, EInvoiceOption.ZUGFERD]
    needs_excel = invoice_option in [EInvoiceOption.EXCEL, EInvoiceOption.PDF, EInvoiceOption.ZUGFERD]
    needs_pdf = invoice_option in [EInvoiceOption.PDF, EInvoiceOption.ZUGFERD]
    needs_zugferd = invoice_option == EInvoiceOption.ZUGFERD

    if needs_excel:
        xls_creator = XLSCreator()
        xls_creator.font_name = FONT_NAME
        ws = xls_creator.workbook.active
        ws.title = "Rechnung"
        xls_creator.set_page_margins(ws, left=2.0, right=0.2, top=1.2, bottom=1.0)
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 7
        ws.column_dimensions["G"].width = 13
        ws.column_dimensions["H"].width = 12
        row = 1
        # logo
        img_path = invoice_data["seller"]["logoData"]
        if img_path and os.path.exists(img_path):
            img = Image(img_path)
            ws.add_image(img, 'H1')
        if needs_zugferd:
            ws.add_image(Image(LOGO_ZUGFERD), f"A{row}")
            xls_creator.set_cell(ws, row, 2, f"invoice by {__title__}", bold=True, italic=True)
        ws.row_dimensions[row].height = 19
        row += 3
        # title
        invoice_title = invoice_data["title"] if invoice_data["title"] else "Rechnung"
        xls_creator.set_cell(ws, row, 1, invoice_title, font_size=24)
        row += 1
        # QR code
        if create_qr_code:
            qr_code = generate_epc_qr(invoice_data, box_size=2, border=1)
            if qr_code:
                row += 2
                xls_creator.set_cell(ws, row, 2, "Überweisen per Code", bold=True, font_size=10, align_vert="bottom")
                row += 1
                xls_creator.set_cell(ws, row, 2, "Ganz bequem mit der Banking-App scannen.", font_size=9, align_vert="top")
                ws.add_image(Image(qr_code), f"A{row - 1}")
                row += 3

    if needs_xml:
        doc = convert_json_to_drafthorse_doc(invoice_data)

    if needs_excel:
        # Header
        section_data = {}
        section_data["Rechnungsnummer"] = invoice_data["number"]  # Rechnungsnummer (BT-1)
        section_data["Rechnungsdatum"] = convert_to_de_date(invoice_data["issueDate"])  # Rechnungsdatum (BT-2)
        section_data["Rechnungstyp"] = create_value_description(invoice_data["typeCode"], INVOICE_TYPE)  # Code für den Rechnungstyp (BT-3)
        section_data["Währung"] = create_value_description(invoice_data["currencyCode"], CURRENCY)  # Code für die Rechnungswährung (BT-5)
        section_data["Fälligkeitsdatum"] = convert_to_de_date(invoice_data["dueDate"])  # Fälligkeitsdatum der Zahlung (BT-9)
        section_data["Leistungs-/Lieferdatum"] = convert_to_de_date(invoice_data["deliveryDate"])  # Tatsächliches Lieferdatum (BT-72)
        if invoice_data["billingPeriodStartDate"]:
            section_data["Leistungs-/Abrechnungszeitraum"] = convert_to_de_date(invoice_data["billingPeriodStartDate"]) + " bis " + \
                convert_to_de_date(invoice_data["billingPeriodEndDate"])  # Rechnungszeitraum (BT-73, BT-74)
        section_data["Käuferreferenz"] = invoice_data["buyerReference"]  # Käuferreferenz (BT-10)
        section_data["Projektnummer"] = invoice_data["projectReference"]  # Projektnummer (BT-11)
        section_data["Vertragsnummer"] = invoice_data["contractReference"]  # Vertragsnummer (BT-12)
        section_data["Bestellnummer"] = invoice_data["purchaseOrderReference"]  # Bestellnummer (BT-13)
        section_data["Auftragsnummer"] = invoice_data["salesOrderReference"]  # Auftragsnummer (BT-14)
        receiving_advice_reference = invoice_data.get("receivingAdviceReference", {})  # Wareneingangsmeldung (BT-15)
        if receiving_advice_reference:
            if receiving_advice_reference["id"]:
                text = receiving_advice_reference["id"]
                date = convert_to_de_date(receiving_advice_reference["issueDate"])
                section_data["Wareneingangsmeldung"] = f"{text} ({date})"
        despatch_advice_reference = invoice_data.get("despatchAdviceReference", {})  # Versandanzeige (BT-16)
        if despatch_advice_reference:
            if despatch_advice_reference["id"]:  # write only date if reference exists
                text = despatch_advice_reference["id"]
                date = convert_to_de_date(despatch_advice_reference["issueDate"])
                section_data["Versandanzeige"] = f"{text} ({date})"
        tender_references = invoice_data.get("tenderReferences", [])  # Ausschreibung/Los (BT-17)
        if tender_references:
            for tender_reference in tender_references:
                section_data["Ausschreibung/Los"] = tender_reference["id"]
                break  # only one in dict
        object_references = invoice_data.get("objectReferences", [])  # Objektreferenz (BT-18)
        if object_references:
            for object_reference in object_references:
                section_data["Objektreferenz"] = object_reference["id"]
                break  # only one in dict
        buyer_accounting_accounts = invoice_data.get("buyerAccountingAccounts", [])  # Buchungskonto des Käufers (BT-19)
        if buyer_accounting_accounts:
            for buyer_accounting_account in buyer_accounting_accounts:
                text = buyer_accounting_account["id"]
                if text:
                    section_data["Buchungskonto des Käufers"] = text
                break  # only one in dict
        invoice_references = invoice_data.get("invoiceReferences", [])  # Rechnungsreferenz (BT-25, BT-26)
        if invoice_references:
            for invoice_reference in invoice_references:
                if invoice_reference["id"]:
                    text = invoice_reference["id"]
                    date = convert_to_de_date(invoice_reference["issueDate"])
                    section_data["Rechnungsreferenz"] = f"{text} ({date})"
                break  # only one in dict
        section_data["Bemerkung"] = invoice_data["note"]  # Bemerkung (BT-22)
        section_data["Einleitungstext"] = invoice_data["introText"]  # Einleitungstext
        row = write_data_to_excel(xls_creator, ws, row, section_data, "Rechnungsdaten")

        # Seller
        section_data = {}
        seller_data = invoice_data["seller"]
        section_data["Unternehmen"] = seller_data["name"]  # Unternehmen (BT-27)
        section_data["Handelsname"] = seller_data["tradeName"]  # Handelsname (BT-28)
        section_data["Verkäuferkennung"] = seller_data["id"]  # Verkäuferkennung (BT-29)
        section_data["Registernummer"] = seller_data["tradeId"]  # Registernummer (BT-30)
        section_data["Umsatzsteuer-ID"] = seller_data["vatId"]  # Umsatzsteuer-ID (BT-31)
        section_data["Steuernummer"] = seller_data["taxId"]  # Steuernummer (BT-32)
        section_data["WEE-Nummer"] = seller_data["weeeId"]  # WEEE-Nummer
        section_data["Rechtliche Informationen"] = seller_data["legalInfo"]  # Rechtliche Informationen (BT-33)
        section_data["Elektronische Adresse"] = seller_data["electronicAddress"]  # Elektronische Adresse (BT-34)
        section_data["Webseite"] = seller_data["websiteText"]  # Webseite Text
        # Anschrift
        seller_address = seller_data["address"]
        section_data["Anschrift"] = SECTION_KEY
        section_data["Straße 1"] = seller_address["line1"]  # Straße 1 (BT-35)
        section_data["Straße 2"] = seller_address["line2"]  # Straße 2 (BT-36)
        section_data["PLZ"] = seller_address["postCode"]  # PLZ (BT-38)
        section_data["Ort"] = seller_address["city"]  # Ort (BT-37)
        section_data["Land"] = create_value_description(seller_address["countryCode"], COUNTRY_CODE)  # Land (BT-40) COUNTRY_CODE
        # Kontakt
        seller_contact = seller_data["contact"]
        section_data["Kontakt"] = SECTION_KEY
        section_data["Name"] = seller_contact["name"]  # Name (BT-41)
        section_data["E-Mail"] = seller_contact["email"]  # E-Mail (BT-43)
        section_data["Telefon"] = seller_contact["phone"]  # Telefon (BT-42)
        section_data["Fax"] = seller_contact["fax"]  # Fax
        row = write_data_to_excel(xls_creator, ws, row, section_data, "Rechnungssteller")

        # Buyer
        section_data = {}
        buyer_data = invoice_data["buyer"]
        section_data["Unternehmen"] = buyer_data["name"]  # Unternehmen (BT-44)
        section_data["Handelsname"] = buyer_data["tradeName"]  # Handelsname (BT-45)
        section_data["Käuferkennung"] = buyer_data["id"]  # Käuferkennung (BT-46)
        section_data["Registernummer"] = buyer_data["tradeId"]  # Registernummer (BT-47)
        section_data["Umsatzsteuer-ID"] = buyer_data["vatId"]  # Umsatzsteuer-ID (BT-48)
        section_data["Elektronische Adresse"] = buyer_data["electronicAddress"]  # Elektronische Adresse (BT-49)
        # Anschrift
        buyer_address = buyer_data["address"]
        section_data["Anschrift"] = SECTION_KEY
        section_data["Straße 1"] = buyer_address["line1"]  # Straße 1 (BT-50)
        section_data["Straße 2"] = buyer_address["line2"]  # Straße 2 (BT-51)
        section_data["PLZ"] = buyer_address["postCode"]  # PLZ (BT-53)
        section_data["Ort"] = buyer_address["city"]  # Ort (BT-52)
        section_data["Land"] = create_value_description(buyer_address["countryCode"], COUNTRY_CODE)  # Land (BT-55) COUNTRY_CODE
        # Kontakt
        buyer_contact = buyer_data["contact"]
        section_data["Kontakt"] = SECTION_KEY
        section_data["Name"] = buyer_contact["name"]  # Name (BT-56)
        section_data["E-Mail"] = buyer_contact["email"]  # E-Mail (BT-58)
        section_data["Telefon"] = buyer_contact["phone"]  # Telefon (BT-57)
        row = write_data_to_excel(xls_creator, ws, row, section_data, "Rechnungsempfänger")

        # Zahlungsdetails
        section_data = {}
        data_payment = invoice_data["payment"]
        if data_payment["methods"]:
            payment_method = data_payment["methods"][0]
            section_data["Zahlungsart"] = create_value_description(payment_method["typeCode"], PAYMENT_METHOD)  # Code für die Zahlungsart PAYMENT_METHOD (BT-81)
            section_data["Kontoinhaber"] = payment_method["accountName"]  # Name des Zahlungskontos (BT-85)
            section_data["IBAN"] = payment_method["iban"]  # Kennung des Zahlungskontos (BT-84)
            section_data["BIC"] = payment_method["bic"]  # Kennung des Zahlungsdienstleisters (BT-86)
            section_data["Name der Bank"] = payment_method["bankName"]  # Name der Bank
        section_data["Verwendungszweck"] = data_payment["reference"]  # Verwendungszweck (BT-83)
        section_data["Zahlungsbedienungen"] = data_payment["terms"]  # Zahlungsbedingungen (BT-20)
        row = write_data_to_excel(xls_creator, ws, row, section_data, "Zahlungsdetails")

        # Lieferdetails
        if "delivery" in invoice_data:
            section_data = {}
            data_delivery = invoice_data["delivery"]
            section_data["Name des Empfängers"] = data_delivery["recipientName"]  # Name des Empfängers (BT-70)
            section_data["Kennung des Lieferorts"] = data_delivery["locationId"]  # Kennung des Lieferorts (BT-71)
            section_data["Lieferadresse"] = SECTION_KEY
            data_delivery_address = data_delivery["address"]
            section_data["Straße 1"] = data_delivery_address["line1"]  # Straße 1 (BT-75)
            section_data["Straße 2"] = data_delivery_address["line2"]  # Straße 2 (BT-76)
            section_data["Zusatz"] = data_delivery_address["line3"]  # Zusatz (BT-165)
            section_data["PLZ"] = data_delivery_address["postCode"]  # PLZ (BT-78)
            section_data["Ort"] = data_delivery_address["city"]  # Ort (BT-77)
            if data_delivery_address["city"]:
                section_data["Land"] = create_value_description(data_delivery_address["countryCode"], COUNTRY_CODE)  # Land (BT-80)
            section_data["Region"] = data_delivery_address["region"]  # Region (BT-79)
            row = write_data_to_excel(xls_creator, ws, row, section_data, "Lieferdetails")

        # Positionen
        for item_number, data_item in enumerate(invoice_data["items"], start=1):
            section_data = {}
            section_data["Name"] = data_item["name"]  # Name (BT-153)
            section_data["Artikel-Nr."] = data_item["id"]  # Artikel-Nr. (BT-155)
            section_data["Steuersatz"] = convert_to_rate(data_item["vatRate"])  # Steuersatz (BT-152)
            section_data["Steuerkategorie"] = create_value_description(data_item["vatCode"], VAT_CODE)  # Steuerkategorie (BT-151)
            if data_item["billingPeriodStart"]:
                section_data["Startdatum"] = convert_to_de_date(data_item["billingPeriodStart"])  # Startdatum (BT-134)
            if data_item["billingPeriodEnd"]:
                section_data["Enddatum"] = convert_to_de_date(data_item["billingPeriodEnd"])  # Enddatum (BT-135)
            section_data["Auftragsposition"] = data_item["orderPosition"]  # Auftragsposition (BT-132)
            object_references = data_item.get("objectReferences", [])
            if object_references:
                for object_reference in object_references:
                    section_data["Objektreferenz"] = object_reference.get("id", "")  # Objektreferenz (BT-128)
                    break
            section_data["Beschreibung"] = data_item["description"]  # Beschreibung (BT-154)
            section_data["Menge"] = data_item["quantity"]  # Menge (BT-129)
            section_data["Einheit"] = create_value_description(data_item["quantityUnit"], UNIT)  # Einheit (BT-130) UNIT
            section_data["Einzelpreis"] = convert_to_de_amount(data_item["netUnitPrice"])  # Einzelpreis (Netto) (BT-146)
            if data_item["basisQuantity"] != 1:
                section_data["Basismenge"] = data_item["basisQuantity"]  # Basismenge (BT-149)
            section_data["Gesamtpreis (Netto)"] = convert_to_de_amount(data_item["netAmount"])  # Gesamtpreis (Netto) (BT-131)
            row = write_data_to_excel(xls_creator, ws, row, section_data, f"Position {item_number}")

        # Nachlässe
        section_data = {}
        for allowance_data in invoice_data.get("allowances", []):
            section_data["Grundbetrag"] = convert_to_de_amount(allowance_data["basisAmount"])  # Grundbetrag (BT-93)
            section_data["Betrag (Netto)"] = convert_to_de_amount(allowance_data["netAmount"])  # Betrag (Netto) (BT-92)
            section_data["Prozent"] = convert_to_rate(allowance_data["percent"])  # Prozent (BT-94)
            section_data["Grund"] = allowance_data["reason"]  # Grund (BT-97)
            section_data["Code des Grundes"] = allowance_data["reasonCode"]  # Code des Grundes (BT-98)
            section_data["Steuerkategorie"] = allowance_data["vatCode"]  # Steuerkategorie (BT-95)
            section_data["Steuersatz"] = convert_to_rate(allowance_data["vatRate"])  # Steuersatz (BT-96)
            row = write_data_to_excel(xls_creator, ws, row, section_data, "Nachlass")

        # Zuschlag
        section_data = {}
        for charge_data in invoice_data.get("charges", []):
            section_data["Grundbetrag"] = convert_to_de_amount(charge_data["basisAmount"])  # Grundbetrag (BT-100)
            section_data["Betrag (Netto)"] = convert_to_de_amount(charge_data["netAmount"])  # Betrag (Netto) (BT-99)
            section_data["Prozent"] = convert_to_rate(charge_data["percent"])  # Prozent (BT-101)
            section_data["Grund"] = charge_data["reason"]  # Grund (BT-104)
            section_data["Code des Grundes"] = charge_data["reasonCode"]  # Code des Grundes (BT-105)
            section_data["Steuerkategorie"] = charge_data["vatCode"]  # Steuerkategorie (BT-102)
            section_data["Steuersatz"] = convert_to_rate(charge_data["vatRate"])  # Steuersatz (BT-103)
            row = write_data_to_excel(xls_creator, ws, row, section_data, "Zuschlag")

        # Steuern
        section_data = {}
        for tax_name, tax_data in invoice_data["taxes"].items():
            section_data["Steuerkategorie"] = create_value_description(tax_data["code"], VAT_CODE)  # Steuerkategorie (BT-118)
            section_data["Steuersatz"] = convert_to_rate(tax_data["rate"])  # Steuersatz (BT-119)
            section_data["Gesamt (Netto)"] = convert_to_de_amount(tax_data["netAmount"])  # Gesamt (Netto) (BT-116)
            section_data["Steuerbetrag"] = convert_to_de_amount(tax_data["vatAmount"])  # Steuerbetrag (BT-117)
            section_data["Grund der Steuerbefreiung"] = tax_data.get("exemptionReason", "")  # Befreiungsgrund (BT-120)
            section_data["Grund der Steuerbefreiung (Code)"] = tax_data.get("exemptionReasonCode", "")  # Code für Befreiungsgrund (BT-121)
            row = write_data_to_excel(xls_creator, ws, row, section_data, f"Steuern {tax_name}")

        # Gesamtsummen
        section_data = {}
        data_totals = invoice_data["totals"]
        section_data["Summe Positionen (Netto)"] = convert_to_de_amount(data_totals["itemsNetAmount"])  # Summe Positionen (Netto) (BT-106)
        section_data["Summe Zuschläge (Netto)"] = convert_to_de_amount(data_totals["chargesNetAmount"])  # Summe Zuschläge (Netto) (BT-108)
        section_data["Summe Nachlässe (Netto)"] = convert_to_de_amount(data_totals["allowancesNetAmount"])  # Summe Nachlässe (Netto) (BT-107)
        section_data["Gesamt (Netto)"] = convert_to_de_amount(data_totals["netAmount"])  # Gesamt (Netto) (BT-109)
        section_data["Summe Umsatzsteuer"] = convert_to_de_amount(data_totals["vatAmount"])  # Summe Umsatzsteuer (BT-110)
        section_data["Gesamt (Brutto)"] = convert_to_de_amount(data_totals["grossAmount"])  # Gesamt (Brutto) (BT-112)
        section_data["Gezahlter Betrag"] = convert_to_de_amount(data_totals["paidAmount"])  # Gezahlter Betrag (BT-113)
        section_data["Rundungsbetrag"] = convert_to_de_amount(data_totals["roundingAmount"])  # Rundungsbetrag (BT-114)
        section_data["Fälliger Betrag"] = convert_to_de_amount(data_totals["dueAmount"])  # Fälliger Betrag (BT-115)
        row = write_data_to_excel(xls_creator, ws, row, section_data, "Gesamtsummen")

    # save file
    if not os.path.exists(EXPORT_PATH):
        os.makedirs(EXPORT_PATH)
    invoice_number = invoice_data["number"]
    buyer_name = invoice_data["buyer"]["name"]  # Unternehmen (BT-44)
    file_description = get_file_name(f"Rechnung_{invoice_number}_{buyer_name}")
    file_name = f"{EXPORT_PATH}/{file_description}"

    valid_xml = True
    if needs_xml:
        file_name_xml = f"{file_name}.xml"
        valid_xml = create_factur_xml(doc, file_name_xml)

    if needs_excel:
        file_name_excel = f"{file_name}.xlsx"
        xls_creator.save(filename=file_name_excel)

    file_to_open = None
    if needs_xml:
        file_to_open = file_name_xml
    if valid_xml:  # continue only if xml is valid or not created
        if needs_excel:
            file_to_open = file_name_excel
        if needs_pdf:
            # Convert Excel to PDF with LibreOffice
            file_name_pdf = f"{file_name}.pdf"
            convert_xlsx_to_pdf(file_name_excel)
            delete_file(file_name_excel)
            file_to_open = file_name_pdf
        if needs_zugferd:
            # Add XML to PDF
            success = add_xml_to_pdf(file_name_pdf, file_name_xml)
            # remove Excel and XML File
            delete_file(file_name_xml)
            if not success:
                QMessageBox.warning(None, "Schreibzugriff", "Rechnung konnte nicht erstellt werden.\nBitte schließen Sie die geöffnete Rechnung.")
                file_to_open = None  # do not open if failed
    else:
        file_to_open = None  # do not open if failed

    return file_to_open


def create_general_invoice(invoice_data: dict[str, Any], invoice_option: EInvoiceOption, create_qr_code: bool) -> None:
    """!
    @brief Create invoice document and open it in the default application.
    @param invoice_data : invoice data as JSON dictionary.
    @param invoice_option : output format option.
    @param create_qr_code : whether to include EPC QR code.
    """
    file = convert_json_to_invoice(invoice_data, invoice_option, create_qr_code)

    if file is not None:
        os.startfile(os.path.abspath(file))
