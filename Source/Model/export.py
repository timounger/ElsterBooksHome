"""!
********************************************************************************
@file   export.py
@brief  Generate financial reports and data exports in XLSX and CSV.
********************************************************************************
"""

import os
import logging
import enum
from typing import Any, TYPE_CHECKING
import copy
import csv
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from Source.version import __title__, __description__, __version__, __copyright__
from Source.Util.app_data import get_computer_name
from Source.Util.openpyxl_util import XLSCreator, NUMBER_FORMAT_EUR, NUMBER_FORMAT_PERCENT, NUMBER_FORMAT_DATETIME, \
    COLOR_YELLOW, COLOR_RED, COLOR_GREEN, COLOR_GREY
from Source.Model.company import LOGO_BRIEF_PATH, ECompanyFields, COMPANY_ADDRESS_FIELD, COMPANY_BOOKING_FIELD
from Source.Model.data_handler import EReceiptFields, EReceiptGroup, MONTHS_IN_YEAR, \
    DATE_FORMAT_JSON, DATE_TIME_FORMAT, MONTH_NAMES_SHORT, is_date_format, calc_vat_rate
if TYPE_CHECKING:
    from Source.Controller.main_window import MainWindow

log = logging.getLogger(__title__)

TOOL_INFO_NAME = "ToolInfo"

UST_REGULATION_DAYS = 10  # 10 days limit for pretax: verrechnete Umsatzsteuer (Die Regelung zum 10- Tageszeitraum nach § 11 Abs. 2 Satz 2 EStG ist zu beachten.)

FONT_SIZE = 11
FONT_NAME = "Calibri"


class EReportType(str, enum.Enum):
    """!
    @brief Available tax report and export types (UST, EÜR, GuV, DATEV).
    """
    UST_PRE = "Umsatzsteuer-Voranmeldung"
    UST = "Umsatzsteuererklärung"
    EUR = "Einnahmenüberschussrechnung"
    GUV = "Gewinn- und Verlustrechnung"
    EXPORT_TOTAL = "Übersicht Gesamt"
    DATEV = "DATEV Export"


class EReportSheet(str, enum.Enum):
    """!
    @brief Worksheet categories within an export report (overview, income, expenditure, etc.).
    """
    OVERVIEW = "Übersicht"
    INCOME = "Einnahmen"
    EXPENDITURE = "Ausgaben"
    PRE_TAX = "Vorsteuer"


class ExportReport:
    """!
    @brief Financial report export generator.
    @param ui : main window object.
    @param report_type : report type.
    @param year : fiscal year.
    @param months : list of months to include.
    @param period : period label for display.
    """

    def __init__(self, ui: "MainWindow", report_type: EReportType, year: int | None = None,
                 months: list[int] | None = None, period: str | None = None) -> None:
        self.ui = ui
        self.report_type = report_type
        self.year = year
        self.months = months
        self.period = period
        self.sheet_overview: Worksheet | None = None
        self.income_sum_row = 1
        self.expenditure_sum_row = 1
        self.pre_tax_sum_row = 1
        self.income_groups: dict[str, int | float] = {}
        self.expenditure_groups: dict[str, int | float] = {}
        self.income_month: dict[int, int | float] = {}
        self.expenditure_month: dict[int, int | float] = {}
        self.income_year: dict[int, int | float] = {}
        self.expenditure_year: dict[int, int | float] = {}
        self.file_name = ""
        self.income_vat_rate_sum_cells: dict[int | float, tuple[int, int]] = {}  # vat_value: (column_number, row_number)
        self.tax_rates = self.ui.tab_settings.company_data[COMPANY_BOOKING_FIELD][ECompanyFields.TAX_RATES]
        self.date_field = self.get_relevant_date_field()

    def get_relevant_date_field(self) -> EReceiptFields:
        """!
        @brief Get relevant date field based on report type.
        @return receipt date field to use for filtering.
        """
        match self.report_type:
            case EReportType.UST | EReportType.UST_PRE:  # for UST is setting in Elster relevant
                agreed_cost = self.ui.tab_settings.company_data[COMPANY_BOOKING_FIELD][ECompanyFields.AGREED_COST]
                if agreed_cost:  # vereinbarte Entgeld
                    field = EReceiptFields.INVOICE_DATE
                else:  # vereinnahmte Entgeld
                    field = EReceiptFields.PAYMENT_DATE
            case EReportType.EUR:  # for EUR is payment date relevant "Zuflussprinzip"
                field = EReceiptFields.PAYMENT_DATE
            case EReportType.GUV:
                field = EReceiptFields.INVOICE_DATE
            case _:  # for other use payment date
                field = EReceiptFields.PAYMENT_DATE
        return field

    def create_xlsx_report(self, file_name: str) -> None:
        """!
        @brief Create xlsx sales report file or CSV for DATEV export.
        @param file_name : name of report file.
        """
        self.file_name = file_name
        xls_creator = XLSCreator(font_name=FONT_NAME, font_size=FONT_SIZE)
        self.create_overview_sheet(xls_creator)
        self.income_sum_row = self.create_receipt_sheet(xls_creator, EReportSheet.INCOME, self.ui.tab_income.receipts, tab_color=COLOR_GREEN)
        self.expenditure_sum_row = self.create_receipt_sheet(xls_creator, EReportSheet.EXPENDITURE, self.ui.tab_expenditure.receipts, tab_color=COLOR_RED)

        if self.report_type != EReportType.DATEV:
            if self.report_type == EReportType.UST:
                self.pre_tax_sum_row = self.create_receipt_sheet(xls_creator, EReportSheet.PRE_TAX, self.ui.tab_expenditure.receipts,
                                                                 invert_records=self.ui.tab_income.receipts, tab_color=COLOR_YELLOW)

            self.add_tax_to_overview_sheet(xls_creator)  # finish overview after other tabs
            self.create_tool_info(xls_creator)

            log.debug("Create File %s", file_name)
            xls_creator.save(filename=file_name)
        else:
            pass  # CSV File already written in function create_receipt_sheet()

    def create_overview_sheet(self, xls_creator: XLSCreator) -> None:
        """!
        @brief Create overview worksheet with company info and period header.
        @param xls_creator : XLS creator instance.
        """
        worksheet = xls_creator.workbook.active
        self.sheet_overview = worksheet
        worksheet.title = EReportSheet.OVERVIEW.value
        worksheet.sheet_properties.tabColor = COLOR_GREY
        xls_creator.set_page_margins(worksheet, left=1.8, right=0.8, top=1.2, bottom=0.8)
        worksheet.column_dimensions["A"].width = 26
        worksheet.column_dimensions["B"].width = 15
        worksheet.column_dimensions["C"].width = 15
        worksheet.column_dimensions["D"].width = 15
        worksheet.column_dimensions["E"].width = 21
        img_path = os.path.join(self.ui.model.data_path, LOGO_BRIEF_PATH)
        if os.path.exists(img_path):
            img = Image(img_path)
            if img.width <= 150:
                worksheet.add_image(img, 'E1')
            else:
                worksheet.add_image(img, 'D1')
        company = self.ui.tab_settings.company_data
        company_address = company[COMPANY_ADDRESS_FIELD]
        xls_creator.set_cell(worksheet, 1, 1, company[ECompanyFields.NAME], italic=True, font_size=14)
        xls_creator.set_cell(worksheet, 2, 1, company_address[ECompanyFields.STREET_1])
        xls_creator.set_cell(worksheet, 3, 1, f"{company_address[ECompanyFields.PLZ]} {company_address[ECompanyFields.CITY]}")
        xls_creator.set_cell(worksheet, 4, 1, f"StNr.: {company[ECompanyFields.TAX_ID]}")
        xls_creator.set_cell(worksheet, 6, 1, self.report_type.value, bold=True, font_size=24)
        title = "Zeitraum:"
        if self.year is not None:
            title += f" {self.year}"
        if self.period is not None:
            title += f" {self.period}"
        if (self.year is None) and (self.period is None):
            title += " Gesamt"
        xls_creator.set_cell(worksheet, 7, 1, title, bold=True)

    def get_relevant_data_status(self, data: dict[EReceiptFields, str], sheet_type: EReportSheet) -> bool:
        """!
        @brief Check if a receipt entry is relevant for the current report.
        @param data : receipt data dictionary.
        @param sheet_type : report sheet type.
        @return True if data is relevant for the report.
        """
        if sheet_type == EReportSheet.PRE_TAX:
            data_relevant = bool(data[EReceiptFields.DESCRIPTION].startswith(f"{EReportType.UST_PRE.value} {self.year}"))
        else:
            date_string = data[self.date_field]
            if date_string and is_date_format(date_string):  # check date format for invalid data
                data_date = datetime.strptime(date_string, DATE_FORMAT_JSON)
                ust_report = bool(self.report_type in [EReportType.UST_PRE, EReportType.UST])
                data_is_ust = bool(data[EReceiptFields.GROUP] in [EReceiptGroup.UST_VA, EReceiptGroup.UST])
                hide_data = bool(ust_report and data_is_ust)
                if self.year is not None:
                    if data_is_ust:
                        if data_date.month == 1 and data_date.day <= UST_REGULATION_DAYS:
                            relevant_year = data_date.year - 1  # relevant for year before
                        else:
                            relevant_year = data_date.year
                    else:
                        relevant_year = data_date.year
                    year_relevant = bool(relevant_year == self.year)
                else:
                    year_relevant = True
                if self.months is not None:
                    month_relevant = bool(data_date.month in self.months)
                else:
                    month_relevant = True
                data_relevant = (not hide_data) and year_relevant and month_relevant
            else:
                data_relevant = False
        return data_relevant

    def create_receipt_sheet(self, xls_creator: XLSCreator, sheet_type: EReportSheet, receipt_data: list[dict[EReceiptFields, Any]],
                             invert_records: list[dict[EReceiptFields, Any]] | None = None, tab_color: str | None = None) -> int:
        """!
        @brief Create receipt worksheet with data rows and summary.
        @param xls_creator : XLS creator instance.
        @param sheet_type : report sheet type.
        @param receipt_data : list of receipt data dictionaries.
        @param invert_records : inverted data for pre-tax calculation.
        @param tab_color : worksheet tab color.
        @return row number of the sum line.
        """
        sheet_title = sheet_type.value
        worksheet = xls_creator.workbook.create_sheet(sheet_title)
        worksheet.title = sheet_title
        if tab_color is not None:
            worksheet.sheet_properties.tabColor = tab_color
        Worksheet.set_printer_settings(worksheet, paper_size=9, orientation="landscape")
        xls_creator.set_page_margins(worksheet, left=1.5, right=1.5, top=1.0, bottom=1.0)
        worksheet.column_dimensions["A"].width = 10  # INVOICE_DATE
        worksheet.column_dimensions["B"].width = 10  # INVOICE_NUMBER
        worksheet.column_dimensions["C"].width = 20  # TRADE_PARTNER
        worksheet.column_dimensions["D"].width = 20  # DESCRIPTION
        worksheet.column_dimensions["E"].width = 10  # PAYMENT_DATE
        worksheet.column_dimensions["F"].width = 4  # BAR
        worksheet.column_dimensions["G"].width = 10  # GROUP
        worksheet.column_dimensions["H"].width = 17  # COMMENT
        worksheet.column_dimensions["I"].width = 17  # DELIVER_DATE
        worksheet.column_dimensions["J"].width = 14  # AMOUNT_GROSS
        worksheet.column_dimensions["K"].width = 14  # AMOUNT_NET
        worksheet.column_dimensions["L"].width = 14  # USt (calculated)
        worksheet.column_dimensions["M"].width = 6  # % USt (calculated)
        worksheet.column_dimensions["N"].width = 36  # ID

        if self.report_type != EReportType.EXPORT_TOTAL:
            worksheet.column_dimensions["H"].hidden = True  # COMMENT
            worksheet.column_dimensions["I"].hidden = True  # DELIVER_DATE
            worksheet.column_dimensions["N"].hidden = True  # ID

        datev_rows = [["Datum", "Belegnummer", "Handelspartner", "Beschreibung", "Bezahlt", "Kategorie", "Netto", "Brutto"]]

        xls_creator.set_cell(worksheet, 1, 1, sheet_title, bold=True, font_size=28)

        records = copy.deepcopy(receipt_data)
        if invert_records is not None:  # only for pretax not None
            for invert_data in copy.deepcopy(invert_records):
                if not isinstance(invert_data[EReceiptFields.AMOUNT_GROSS], (int, float)) or not isinstance(invert_data[EReceiptFields.AMOUNT_NET], (int, float)):
                    raise TypeError(
                        f"Invalid amount type in receipt '{invert_data[EReceiptFields.ID]}': gross={invert_data[EReceiptFields.AMOUNT_GROSS]!r}, net={invert_data[EReceiptFields.AMOUNT_NET]!r}")
                invert_data[EReceiptFields.AMOUNT_GROSS] = -invert_data[EReceiptFields.AMOUNT_GROSS]
                invert_data[EReceiptFields.AMOUNT_NET] = -invert_data[EReceiptFields.AMOUNT_NET]
                records.append(invert_data)

        # Get relevant data
        vat_rate_breakdown: dict[float | int, float | int] = {}
        for tax_rate in self.tax_rates:
            vat_rate_breakdown[tax_rate] = 0
        relevant_data = []
        for entry in records:
            if (self.report_type == EReportType.EXPORT_TOTAL) or self.get_relevant_data_status(entry, sheet_type):
                relevant_data.append(entry)
                # Vat Rate breakdown
                vat_rate = calc_vat_rate(entry[EReceiptFields.AMOUNT_GROSS], entry[EReceiptFields.AMOUNT_NET])
                vat_rate_breakdown[vat_rate] = 0  # create empty vat to know present rates for sort
        vat_rate_breakdown_sorted = dict(sorted(vat_rate_breakdown.items(), key=lambda item: item[0], reverse=True))

        receipt_idx = 0
        row = receipt_idx
        for entry in relevant_data:
            row = receipt_idx + 3
            xls_creator.set_cell(worksheet, row, 1, entry[EReceiptFields.INVOICE_DATE], number_format=NUMBER_FORMAT_DATETIME)
            xls_creator.set_cell(worksheet, row, 2, entry[EReceiptFields.INVOICE_NUMBER])
            xls_creator.set_cell(worksheet, row, 3, entry[EReceiptFields.TRADE_PARTNER])
            xls_creator.set_cell(worksheet, row, 4, entry[EReceiptFields.DESCRIPTION])
            xls_creator.set_cell(worksheet, row, 5, entry[EReceiptFields.PAYMENT_DATE], number_format=NUMBER_FORMAT_DATETIME)
            if entry[EReceiptFields.BAR]:
                xls_creator.set_cell(worksheet, row, 6, "X")
            xls_creator.set_cell(worksheet, row, 7, entry[EReceiptFields.GROUP])
            xls_creator.set_cell(worksheet, row, 8, entry[EReceiptFields.COMMENT])
            xls_creator.set_cell(worksheet, row, 9, entry[EReceiptFields.DELIVER_DATE])
            xls_creator.set_cell(worksheet, row, 10, entry[EReceiptFields.AMOUNT_GROSS], number_format=NUMBER_FORMAT_EUR)
            xls_creator.set_cell(worksheet, row, 11, entry[EReceiptFields.AMOUNT_NET], number_format=NUMBER_FORMAT_EUR)
            xls_creator.set_cell(worksheet, row, 12, f"=(J{row}-K{row})", number_format=NUMBER_FORMAT_EUR)
            xls_creator.set_cell(worksheet, row, 13, f"=(IF(K{row}=0, 0, (J{row}/K{row})-1))", number_format=NUMBER_FORMAT_PERCENT)
            xls_creator.set_cell(worksheet, row, 14, entry[EReceiptFields.ID])
            # Vat Rate breakdown
            if sheet_type in [EReportSheet.INCOME, EReportSheet.EXPENDITURE]:
                vat_rate = calc_vat_rate(entry[EReceiptFields.AMOUNT_GROSS], entry[EReceiptFields.AMOUNT_NET])
                tax_amount = entry[EReceiptFields.AMOUNT_GROSS] - entry[EReceiptFields.AMOUNT_NET]
                vat_rate_breakdown_sorted[vat_rate] = vat_rate_breakdown_sorted.get(vat_rate, 0) + entry[EReceiptFields.AMOUNT_NET]
                vat_column = 15 + (2 * list(vat_rate_breakdown_sorted.keys()).index(vat_rate))
                xls_creator.set_cell(worksheet, row, vat_column, entry[EReceiptFields.AMOUNT_NET], number_format=NUMBER_FORMAT_EUR)
                xls_creator.set_cell(worksheet, row, vat_column + 1, tax_amount, number_format=NUMBER_FORMAT_EUR)
            receipt_idx += 1
            # add group data for eur
            if self.report_type in [EReportType.EUR, EReportType.GUV, EReportType.EXPORT_TOTAL]:
                if entry[self.date_field]:  # missing data for total export possible
                    group = entry[EReceiptFields.GROUP]
                    data_date = datetime.strptime(entry[self.date_field], DATE_FORMAT_JSON)
                    month = data_date.month if (data_date.year == self.year) else 12
                    data_is_ust = bool(entry[EReceiptFields.GROUP] in [EReceiptGroup.UST_VA, EReceiptGroup.UST])
                    year = data_date.year - 1 if (data_is_ust and (data_date.month == 1) and (data_date.day <= UST_REGULATION_DAYS)) else data_date.year
                    match sheet_type:
                        case EReportSheet.INCOME:
                            if self.report_type in [EReportType.EUR, EReportType.GUV]:
                                self.income_groups[group] = self.income_groups.get(group, 0) + entry[EReceiptFields.AMOUNT_NET]
                                self.income_month[month] = self.income_month.get(month, 0) + entry[EReceiptFields.AMOUNT_GROSS]
                            if self.report_type == EReportType.EXPORT_TOTAL:
                                self.income_year[year] = self.income_year.get(year, 0) + entry[EReceiptFields.AMOUNT_GROSS]
                        case EReportSheet.EXPENDITURE:
                            if self.report_type in [EReportType.EUR, EReportType.GUV]:
                                self.expenditure_groups[group] = self.expenditure_groups.get(group, 0) + entry[EReceiptFields.AMOUNT_NET]
                                self.expenditure_month[month] = self.expenditure_month.get(month, 0) + entry[EReceiptFields.AMOUNT_GROSS]
                            if self.report_type == EReportType.EXPORT_TOTAL:
                                self.expenditure_year[year] = self.expenditure_year.get(year, 0) + entry[EReceiptFields.AMOUNT_GROSS]
            # set data for CSV export
            if self.report_type == EReportType.DATEV:
                # use only first word of group if is number
                group = entry[EReceiptFields.GROUP]
                words = group.split()
                if words and words[0].isdigit():
                    group = words[0]
                row_data = [
                    entry[EReceiptFields.INVOICE_DATE],  # Datum
                    entry[EReceiptFields.INVOICE_NUMBER],  # Belegnummer
                    entry[EReceiptFields.TRADE_PARTNER],  # Handelspartner
                    entry[EReceiptFields.DESCRIPTION],  # Beschreibung
                    entry[EReceiptFields.PAYMENT_DATE],  # Bezahlt
                    group,  # Kategorie
                    str(entry[EReceiptFields.AMOUNT_NET]).replace(".", ","),  # Netto
                    str(entry[EReceiptFields.AMOUNT_GROSS]).replace(".", ","),  # Brutto
                ]
                datev_rows.append(row_data)

        # set title after iterate over entry for known column because vat rate breakdown
        header = ["Datum", "Re-Nr.", "Handelspartner", "Beschreibung", "Bezahlt", "Bar", "Gruppe", "Kommentar", "Lieferzeitraum", "Brutto", "Netto", "USt.", "% USt.", "ID"]
        if sheet_type in [EReportSheet.INCOME, EReportSheet.EXPENDITURE]:
            group_letter_start = get_column_letter(len(header) + 1)
            for vat_rate in vat_rate_breakdown_sorted:
                header_row = len(header) + 1
                header_letter = get_column_letter(header_row)
                header_row_next = header_row + 1
                header_letter_next = get_column_letter(header_row_next)
                worksheet.column_dimensions[header_letter].width = 20
                worksheet.column_dimensions[header_letter_next].width = 20
                worksheet.merge_cells(f"{header_letter}1:{header_letter_next}1")
                xls_creator.set_cell(worksheet, 1, header_row, f"Steuersatz {vat_rate}%", bold=True, italic=True, align="center", align_vert="bottom")
                header += [f"Grundbetrag zu {vat_rate}%", f"Steuerbetrag zu {vat_rate}%"]
                if sheet_type == EReportSheet.INCOME:
                    self.income_vat_rate_sum_cells[vat_rate] = (header_row, row + 1)
            if vat_rate_breakdown_sorted:
                worksheet.column_dimensions.group(group_letter_start, header_letter_next, hidden=True)
        for i, title in enumerate(header):
            xls_creator.set_cell(worksheet, 2, i + 1, title, bold=True)

        if self.report_type == EReportType.DATEV:
            datev_file_name = None
            export_dir = os.path.dirname(self.file_name)
            match sheet_type:
                case EReportSheet.INCOME:
                    datev_file_name = f"{export_dir}/Rechnungseingangsbelege.csv"
                case EReportSheet.EXPENDITURE:
                    datev_file_name = f"{export_dir}/Rechnungsausgangsbelege.csv"
            if datev_file_name is not None:
                with open(datev_file_name, mode="w", encoding="utf-8-sig", newline="") as file:  # utf-8-sig for umlauts
                    writer = csv.writer(file, delimiter=";")
                    for datev_entry in datev_rows:
                        writer.writerow(datev_entry)
        if receipt_idx == 0:  # set correct row if no data
            receipt_idx = 1
            row = receipt_idx + 2
        xls_creator.set_table(worksheet, max_col=len(header), max_row=receipt_idx + 2, min_col=1, min_row=2)

        row += 1
        sum_header = [""] * len(header)
        sum_header[0] = "Summe"
        sum_header[9] = f"=SUM({sheet_title}[Brutto])"
        sum_header[10] = f"=SUM({sheet_title}[Netto])"
        sum_header[11] = f"=SUM({sheet_title}[USt.])"
        if len(sum_header) > 14:
            for i, vat_rate in enumerate(vat_rate_breakdown_sorted):
                sum_header[14 + (i * 2)] = f"=SUM({sheet_title}[Grundbetrag zu {vat_rate}%])"
                sum_header[14 + (i * 2) + 1] = f"=SUM({sheet_title}[Steuerbetrag zu {vat_rate}%])"
        for i, title in enumerate(sum_header):
            xls_creator.set_cell(worksheet, row, i + 1, title, bold=True, number_format=NUMBER_FORMAT_EUR, fill_color=COLOR_YELLOW)

        return row

    def add_tax_to_overview_sheet(self, xls_creator: XLSCreator) -> None:
        """!
        @brief Add tax data to overview sheet.
        @param xls_creator : XLS creator instance.
        """
        assert self.sheet_overview is not None
        worksheet = self.sheet_overview
        row = 9
        match self.report_type:
            case EReportType.UST_PRE | EReportType.UST:
                sum_row = 5
                sum_row_letter = get_column_letter(sum_row)
                xls_creator.set_cell(worksheet, row, 1, "Art der Besteuerung:")
                agreed_cost = self.ui.tab_settings.company_data[COMPANY_BOOKING_FIELD][ECompanyFields.AGREED_COST]
                tax_method = "Soll-Versteuerung (bei Rechnungsstellung)" if agreed_cost else "Ist-Versteuerung (bei Eingang der Zahlung)"
                xls_creator.set_cell(worksheet, row, 2, tax_method)
                row += 2
                # income
                xls_creator.set_cell(worksheet, row, 1, "Steuerpflichtige Umsätze")
                xls_creator.set_cell(worksheet, row, sum_row - 2, "Bemessungsgrundlage")
                xls_creator.set_cell(worksheet, row, sum_row, "Steuer")
                tax_sum_row_start = row + 1
                for vat_rate, cell_position in self.income_vat_rate_sum_cells.items():
                    column_number, row_number = cell_position
                    row += 1
                    tax_sum_row_end = row
                    net_value_cell = f"{get_column_letter(column_number)}{row_number}"
                    tax_value_cell = f"{get_column_letter(column_number + 1)}{row_number}"
                    xls_creator.set_cell(worksheet, row, 1, f"zum Steuersatz von {vat_rate} Prozent")
                    xls_creator.set_cell(worksheet, row, sum_row - 1, f"={EReportSheet.INCOME.value}!{net_value_cell}", number_format=NUMBER_FORMAT_EUR)
                    xls_creator.set_cell(worksheet, row, sum_row, f"={EReportSheet.INCOME.value}!{tax_value_cell}", number_format=NUMBER_FORMAT_EUR)
                # out tax
                row += 2
                xls_creator.set_cell(worksheet, row, 1, "Abziehbare Vorsteuerbeträge")
                row += 1
                pre_tax_other_row = row
                xls_creator.set_cell(worksheet, row, 1, "Vorsteuerbeträge aus Rechnungen von anderen Unternehmern")
                xls_creator.set_cell(worksheet, row, sum_row, f"={EReportSheet.EXPENDITURE.value}!L{self.expenditure_sum_row}", number_format=NUMBER_FORMAT_EUR)
                # pre pay
                tax_name = "Umsatzsteuer-Vorauszahlung / Überschuss (Steuer)" if (self.report_type == EReportType.UST_PRE) else "Umsatzsteuer"
                row += 2
                ust_row = row
                xls_creator.set_cell(worksheet, row, 1, tax_name)
                xls_creator.set_cell(worksheet, row, sum_row,
                                     f"=SUM({sum_row_letter}{tax_sum_row_start}:{sum_row_letter}{tax_sum_row_end})-{sum_row_letter}{pre_tax_other_row}",
                                     number_format=NUMBER_FORMAT_EUR)
                if self.report_type == EReportType.UST:
                    row += 2
                    ust_already_row = row
                    xls_creator.set_cell(worksheet, row, 1, "Bereits entrichtete Vorsteuerbeträge")
                    xls_creator.set_cell(worksheet, row, sum_row, f"={EReportSheet.PRE_TAX.value}!J{self.pre_tax_sum_row}", number_format=NUMBER_FORMAT_EUR)
                    row += 2
                    xls_creator.set_cell(worksheet, row, 1, "Noch an die Finanzkasse zu entrichten")
                    xls_creator.set_cell(worksheet, row, sum_row, f"=({sum_row_letter}{ust_row}-{sum_row_letter}{ust_already_row})", number_format=NUMBER_FORMAT_EUR)
            case EReportType.EUR | EReportType.GUV:
                sum_row = 3
                sum_row_letter = get_column_letter(sum_row)
                special_groups = [EReceiptGroup.UST_VA, EReceiptGroup.UST]  # write special groups at end
                # income
                xls_creator.set_cell(worksheet, row, 1, "1. Betriebseinnahmen (einschl. steuerfreier Betriebseinnahmen)", bold=True)
                row += 1
                start_row = row
                income_groups_sorted = dict(sorted(self.income_groups.items(), key=lambda item: (item[0] in special_groups, item[0].lower())))
                for income_group, income_net in income_groups_sorted.items():
                    xls_creator.set_cell(worksheet, row, 1, income_group)
                    xls_creator.set_cell(worksheet, row, sum_row, income_net, number_format=NUMBER_FORMAT_EUR)
                    row += 1
                xls_creator.set_cell(worksheet, row, 1, "Vereinnahmte Umsatzsteuer")
                xls_creator.set_cell(worksheet, row, sum_row, f"={EReportSheet.INCOME.value}!L{self.income_sum_row}", number_format=NUMBER_FORMAT_EUR)
                row += 1
                income_sum_overview_row = row
                xls_creator.set_cell(worksheet, row, 1, "Summe", bold=True)
                xls_creator.set_cell(worksheet, row, sum_row, f"=SUM({sum_row_letter}{start_row}:{sum_row_letter}{row - 1})", bold=True, number_format=NUMBER_FORMAT_EUR)
                # expenditure
                row += 2
                xls_creator.set_cell(worksheet, row, 1, "2. Betriebsausgaben (einschl. auf steuerfreie Betriebseinnahmen entfallende Betriebsausgaben)", bold=True)
                row += 1
                start_row = row
                expenditure_groups_sorted = dict(sorted(self.expenditure_groups.items(), key=lambda item: (item[0] in special_groups, item[0].lower())))
                for expenditure_group, expenditure_net in expenditure_groups_sorted.items():
                    xls_creator.set_cell(worksheet, row, 1, expenditure_group)
                    xls_creator.set_cell(worksheet, row, sum_row, expenditure_net, number_format=NUMBER_FORMAT_EUR)
                    row += 1
                xls_creator.set_cell(worksheet, row, 1, "Gezahlte Vorsteuerbeträge")
                xls_creator.set_cell(worksheet, row, sum_row, f"={EReportSheet.EXPENDITURE.value}!L{self.expenditure_sum_row}", number_format=NUMBER_FORMAT_EUR)
                row += 1
                expenditure_sum_overview_row = row
                xls_creator.set_cell(worksheet, row, 1, "Summe", bold=True)
                xls_creator.set_cell(worksheet, row, sum_row, f"=SUM({sum_row_letter}{start_row}:{sum_row_letter}{row - 1})", bold=True, number_format=NUMBER_FORMAT_EUR)
                # profit
                row += 2
                xls_creator.set_cell(worksheet, row, 1, "3. Ermittlung des Gewinns", bold=True)
                row += 1
                xls_creator.set_cell(worksheet, row, 1, "Summe", bold=True)
                xls_creator.set_cell(worksheet, row, sum_row, f"={sum_row_letter}{income_sum_overview_row}-{sum_row_letter}{expenditure_sum_overview_row}", bold=True, number_format=NUMBER_FORMAT_EUR)

                # create chart with monthly data
                row += 3
                xls_creator.set_cell(worksheet, row, 1, "Monat")
                xls_creator.set_cell(worksheet, row, 2, "Einnahmen")
                xls_creator.set_cell(worksheet, row, 3, "Ausgaben")
                xls_creator.set_cell(worksheet, row, 4, "Differenz")
                row += 1
                for month in range(MONTHS_IN_YEAR):
                    xls_creator.set_cell(worksheet, row + month, 1, MONTH_NAMES_SHORT[month])
                    income_month_sum = self.income_month.get(month + 1, 0)
                    xls_creator.set_cell(worksheet, row + month, 2, income_month_sum, number_format=NUMBER_FORMAT_EUR)
                    expenditure_month_sum = - self.expenditure_month.get(month + 1, 0)  # use negative for diagram view
                    xls_creator.set_cell(worksheet, row + month, 3, expenditure_month_sum, number_format=NUMBER_FORMAT_EUR)
                    xls_creator.set_cell(worksheet, row + month, 4, f"=B{row + month}+C{row + month}", number_format=NUMBER_FORMAT_EUR)

                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = self.report_type.value
                if chart.x_axis is not None:
                    chart.x_axis.title = "Monat"
                if chart.y_axis is not None:
                    chart.y_axis.title = "Summe"
                chart.height = 10
                chart.width = 17.5
                chart.legend = None

                month_data = Reference(worksheet, min_col=1, min_row=row, max_row=row + 11, max_col=1)

                income_data = Reference(worksheet, min_col=2, min_row=row, max_row=row + 11, max_col=2)
                chart.add_data(income_data, titles_from_data=False)

                expenditure_data = Reference(worksheet, min_col=3, min_row=row, max_row=row + 11, max_col=3)
                chart.add_data(expenditure_data, titles_from_data=False)

                # set color
                income_series = chart.series[0]
                income_series.graphicalProperties.solidFill = COLOR_GREEN
                expenditure_series = chart.series[1]
                expenditure_series.graphicalProperties.solidFill = COLOR_RED

                chart.set_categories(month_data)
                chart.shape = 4
                worksheet.add_chart(chart, f"A{row + 12}")
            case _:
                xls_creator.set_cell(worksheet, row, 2, "Brutto")
                xls_creator.set_cell(worksheet, row, 3, "Netto")
                xls_creator.set_cell(worksheet, row, 4, "Steuer")
                row += 1
                income_row = row
                xls_creator.set_cell(worksheet, row, 1, EReportSheet.INCOME.value)
                xls_creator.set_cell(worksheet, row, 2, f"={EReportSheet.INCOME.value}!J{self.income_sum_row}", number_format=NUMBER_FORMAT_EUR)
                xls_creator.set_cell(worksheet, row, 3, f"={EReportSheet.INCOME.value}!K{self.income_sum_row}", number_format=NUMBER_FORMAT_EUR)
                xls_creator.set_cell(worksheet, row, 4, f"={EReportSheet.INCOME.value}!L{self.income_sum_row}", number_format=NUMBER_FORMAT_EUR)
                row += 1
                expenditure_row = row
                xls_creator.set_cell(worksheet, row, 1, EReportSheet.EXPENDITURE.value)
                xls_creator.set_cell(worksheet, row, 2, f"={EReportSheet.EXPENDITURE.value}!J{self.expenditure_sum_row}", number_format=NUMBER_FORMAT_EUR)
                xls_creator.set_cell(worksheet, row, 3, f"={EReportSheet.EXPENDITURE.value}!K{self.expenditure_sum_row}", number_format=NUMBER_FORMAT_EUR)
                xls_creator.set_cell(worksheet, row, 4, f"={EReportSheet.EXPENDITURE.value}!L{self.expenditure_sum_row}", number_format=NUMBER_FORMAT_EUR)
                row += 1
                xls_creator.set_cell(worksheet, row, 1, "Differenz", bold=True)
                xls_creator.set_cell(worksheet, row, 2, f"=B{income_row}-B{expenditure_row}", bold=True, number_format=NUMBER_FORMAT_EUR)
                xls_creator.set_cell(worksheet, row, 3, f"=C{income_row}-C{expenditure_row}", bold=True, number_format=NUMBER_FORMAT_EUR)
                xls_creator.set_cell(worksheet, row, 4, f"=D{income_row}-D{expenditure_row}", bold=True, number_format=NUMBER_FORMAT_EUR)

                # create chart with year data
                if self.income_year or self.expenditure_year:
                    row += 3
                    xls_creator.set_cell(worksheet, row, 1, "Jahr")
                    xls_creator.set_cell(worksheet, row, 2, "Einnahmen")
                    xls_creator.set_cell(worksheet, row, 3, "Ausgaben")
                    xls_creator.set_cell(worksheet, row, 4, "Differenz")
                    row += 1
                    income_min_year = min(self.income_year.keys()) if self.income_year else None
                    expenditure_min_year = min(self.expenditure_year.keys()) if self.expenditure_year else None
                    if income_min_year is None:
                        income_min_year = expenditure_min_year
                    if expenditure_min_year is None:
                        expenditure_min_year = income_min_year
                    assert income_min_year is not None and expenditure_min_year is not None
                    income_max_year = max(self.income_year.keys()) if self.income_year else income_min_year
                    expenditure_max_year = max(self.expenditure_year.keys()) if self.expenditure_year else expenditure_min_year
                    min_year = min(income_min_year, expenditure_min_year)
                    max_year = max(income_max_year, expenditure_max_year)
                    year_idx = 0
                    for year in range(min_year, max_year + 1):
                        xls_creator.set_cell(worksheet, row + year_idx, 1, year)
                        income_year_sum = self.income_year.get(year, 0)
                        xls_creator.set_cell(worksheet, row + year_idx, 2, income_year_sum, number_format=NUMBER_FORMAT_EUR)
                        expenditure_year_sum = - self.expenditure_year.get(year, 0)  # use negative for diagram view
                        xls_creator.set_cell(worksheet, row + year_idx, 3, expenditure_year_sum, number_format=NUMBER_FORMAT_EUR)
                        xls_creator.set_cell(worksheet, row + year_idx, 4, f"=B{row + year_idx}+C{row + year_idx}", number_format=NUMBER_FORMAT_EUR)
                        year_idx += 1

                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = self.report_type.value
                    if chart.x_axis is not None:
                        chart.x_axis.title = "Jahr"
                    if chart.y_axis is not None:
                        chart.y_axis.title = "Summe"
                    chart.height = 15
                    chart.width = 30
                    chart.legend = None

                    month_data = Reference(worksheet, min_col=1, min_row=row, max_row=row + year_idx - 1, max_col=1)

                    income_data = Reference(worksheet, min_col=2, min_row=row, max_row=row + year_idx - 1, max_col=2)
                    chart.add_data(income_data, titles_from_data=False)

                    expenditure_data = Reference(worksheet, min_col=3, min_row=row, max_row=row + year_idx - 1, max_col=3)
                    chart.add_data(expenditure_data, titles_from_data=False)

                    # set color
                    income_series = chart.series[0]
                    income_series.graphicalProperties.solidFill = COLOR_GREEN
                    expenditure_series = chart.series[1]
                    expenditure_series.graphicalProperties.solidFill = COLOR_RED

                    chart.set_categories(month_data)
                    chart.shape = 4
                    worksheet.add_chart(chart, f"A{row + year_idx}")

    def create_tool_info(self, xls_creator: XLSCreator) -> None:
        """!
        @brief Write a hidden sheet with application version and creation timestamp.
        @param xls_creator : XLS workbook builder instance.
        """
        ws = xls_creator.workbook.create_sheet(TOOL_INFO_NAME)
        ws.sheet_state = "hidden"
        create_time = datetime.now().strftime(DATE_TIME_FORMAT)
        xls_creator.set_cell(ws, 1, 1, f"This export was generated on {create_time} with:")
        xls_creator.set_cell(ws, 3, 1, __title__, bold=True)
        xls_creator.set_cell(ws, 4, 1, __description__)
        xls_creator.set_cell(ws, 5, 1, f"Version: {__version__}")
        xls_creator.set_cell(ws, 6, 1, __copyright__)
        xls_creator.set_cell(ws, 8, 1, f"Device name: {get_computer_name()}")
        ws.column_dimensions["A"].width = 55
